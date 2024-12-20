import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def add_new_column(file_path: str, column_name: str) -> None:
    """
    Adds a new column with the specified name to an Excel workbook.

    Args:
    file_path (str): Path to the Excel file.
    column_name (str): Name of the new column to be added.

    Returns:
        None
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Select the active sheet
        sheet = workbook.active

        # Determine the next available column index
        next_col_index = sheet.max_column + 1

        # Set the column name in the first row of the new column
        sheet.cell(row=1, column=next_col_index, value=column_name)

        # Save the workbook
        workbook.save(file_path)

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{file_path}' is not a valid Excel file.")