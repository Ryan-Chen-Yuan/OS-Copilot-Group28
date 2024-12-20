import openpyxl
from datetime import datetime

def sort_excel_by_date(file_path: str, sheet_name: str) -> None:
    """
    Sorts the columns of an Excel sheet based on the date in the first row.

    Args:
    file_path (str): Path to the Excel file.
    sheet_name (str): Name of the sheet to be sorted.

    Returns:
    None: The function modifies the file in place and does not return anything.
    """
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Extract the header row (dates)
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=11, values_only=True))[0]

    # Create a list of (date, column_index) tuples
    date_columns = []
    for idx, header in enumerate(header_row, start=2):
        try:
            # Attempt to parse the header as a date
            date = datetime.strptime(str(header), "%Y-%m-%d")
            date_columns.append((date, idx))
        except (ValueError, TypeError):
            # Skip columns that do not have a valid date
            continue

    # Sort columns by date
    date_columns.sort()

    # Create a new sheet to store sorted data
    sorted_sheet = workbook.create_sheet(title=f"{sheet_name}_sorted")

    # Copy the first column (stock names) to the new sheet
    for row_idx, cell in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        sorted_sheet.cell(row=row_idx, column=1, value=cell[0])

    # Write sorted columns to the new sheet
    for new_col_idx, (date, old_col_idx) in enumerate(date_columns, start=2):
        for row_idx, cell in enumerate(sheet.iter_rows(min_col=old_col_idx, max_col=old_col_idx, values_only=True), start=1):
            sorted_sheet.cell(row=row_idx, column=new_col_idx, value=cell[0])

    # Remove the original sheet and rename the sorted sheet
    workbook.remove(sheet)
    sorted_sheet.title = sheet_name

    # Save the workbook
    workbook.save(file_path)

    print("Columns sorted by date successfully!")
