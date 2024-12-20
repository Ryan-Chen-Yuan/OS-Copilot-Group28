import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Any, Optional

def read_stock_data_by_name(file_path: str, stock_name: str, sheet_name: Optional[str] = None) -> List[Any]:
    """
    Reads the row data for a specified stock name from an Excel sheet.

    Args:
    file_path (str): Path to the Excel file.
    stock_name (str): The name of the stock to find in the first column.
    sheet_name (Optional[str]): Name of the sheet to be read. If None, reads from the first sheet.

    Returns:
    List[Any]: A list containing data for the specified stock across 10 weeks.

    """
    try:
        # Load the workbook and the specified sheet
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # If sheet_name is not provided, use the first sheet
        if sheet_name is None:
            sheet = workbook.active
        else:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        # Find the row for the specified stock name
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] == stock_name:
                # Get the data for the stock from the second column onwards
                return list(row[1:11])  # Extracting 10 weeks of data from the second column

        raise ValueError(f"Stock '{stock_name}' not found in the first column.")

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{file_path}' is not a valid Excel file.")