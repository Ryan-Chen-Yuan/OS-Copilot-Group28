import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

def cal_average(file_path: str, result_column: str = "Average") -> None:
    """
    Calculates the average of each row in an Excel workbook and adds the result to a new column.

    Args:
    file_path (str): Path to the Excel file.
    result_column (str): Name of the column where the averages will be stored. Default is "Average".

    Returns:
        None
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Select the active sheet
        sheet = workbook.active

        # Determine the next available column index for storing averages
        next_col_index = sheet.max_column + 1

        # Set the result column name in the first row
        sheet.cell(row=1, column=next_col_index, value=result_column)

        # Iterate over each row to calculate the average
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            # Extract numeric values from the row
            values = [cell.value for cell in row if isinstance(cell.value, (int, float))]
            
            # Calculate the average if there are numeric values
            if values:
                average = sum(values) / len(values)
                # Store the average in the new column
                sheet.cell(row=row[0].row, column=next_col_index, value=average)

        # Save the workbook
        workbook.save(file_path)

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{file_path}' is not a valid Excel file.")
