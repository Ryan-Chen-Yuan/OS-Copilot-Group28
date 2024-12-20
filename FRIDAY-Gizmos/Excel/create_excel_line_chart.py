from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

def create_excel_line_chart(excel_path, chart_title, sheet_name):
    """
    在Excel文件中创建一个包含所有聚类的折线图。

    Args:
    excel_path (str): Excel文件的路径。
    chart_title (str): 折线图的标题。
    sheet_name (str): 工作表的名称。

    Returns:
    None
    """
    # 加载工作簿和工作表
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]

    # 获取聚类信息
    clusters = set(sheet.cell(row=row, column=sheet.max_column).value for row in range(2, sheet.max_row + 1))

    # 创建折线图
    chart = LineChart()
    chart.title = chart_title
    chart.style = 13
    chart.y_axis.title = 'Price Change'
    chart.x_axis.title = 'Date'

    # 定义类别（时间点）
    categories = Reference(sheet, min_col=2, min_row=1, max_col=sheet.max_column - 2, max_row=1)
    chart.set_categories(categories)

    for cluster in clusters:
        # 找到属于该聚类的第一只股票
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=sheet.max_column).value == cluster:
                stock_row = row
                break

        # 定义数据范围
        data = Reference(sheet, min_col=2, min_row=stock_row, max_col=sheet.max_column - 2, max_row=stock_row)

        # 添加数据到图表
        chart.add_data(data, titles_from_data=True, from_rows=True)

    # 将图表添加到工作表
    sheet.add_chart(chart, "B15")

    # 保存工作簿
    wb.save(excel_path)

    print("Line chart is successfully created in Excel!")
